import AStar
import Board as bd
import Node
from AStar import AStar
from GBFS import GBFS
from UCS import UCS
import time
import sys
import collections
from openpyxl import Workbook



directory = "SampleInputOutput/Sample/"
output_directory = "Output/"

file = open(directory + "puzzle-input.txt", 'r')

boards = []

# Identifying the lines in the output
for line in file.readlines():
    if line[0] != '#' and line.strip() != "":
        board = bd.Board(line)
        boards.append(board)

start_time = time.time()
workbook = Workbook()
sheet = workbook.active

sheet['A1'] = "Puzzle Number"
sheet['B1'] = "Algorithm"
sheet['C1'] = "Heuristic"
sheet['D1'] = "Solution Length"
sheet['E1'] = "Search Path Length"
sheet['F1'] = "Run Time"

for index, board in enumerate(boards):
    sheet.cell(row=index*9+1+1, column=1).value = index*9+1
    sheet.cell(row=index*9+1+1, column=2).value = "UCS"
    sheet.cell(row=index*9+1+1, column=3).value = "N/A"
    solution_file = open(output_directory + "ucs" + "-sol-" + str(index + 1), 'w')
    search_file = open(output_directory + "ucs" + "-search-" + str(index + 1), 'w')

    ucs_solver = UCS(board)
    ucs_solver.search_solution()

    if ucs_solver.solutionFound:
        solution_file.write("--------------------------------------------------------------------------------\n\n")
        solution_file.write("Initial Configuration: \n\n")
        solution_file.write(board.returnBoardInfo())
        solution_file.write(board.printCarFuel() + "\n")

        solution_file.write("Runtime: " + str(ucs_solver.runtime) + " seconds\n")
        solution_file.write("Search Path Length: " + str(len(ucs_solver.searchPath)-1) + " states\n")
        solution_file.write("Solution Path Length: " + str(len(ucs_solver.solutionPath) - 1) + " moves\n")
        solution_file.write("Solution Path: " + ucs_solver.get_solution_moves() + "\n\n")
        solution_file.write(ucs_solver.get_board_line() + "\n\n")

        solution_file.write("Final Configuration: \n\n")

        solution_file.write(ucs_solver.solutionPath[0].board.returnBoardInfo())
        solution_file.close()
        sheet.cell(row=index*9+1+1, column=4).value = len(ucs_solver.solutionPath) - 1
        sheet.cell(row=index*9+1+1, column=5).value = len(ucs_solver.searchPath) - 1
        sheet.cell(row=index*9+1+1, column=6).value = ucs_solver.runtime
    else:
        solution_file.write("--------------------------------------------------------------------------------\n\n")
        solution_file.write("Initial Configuration: \n\n")
        solution_file.write(board.returnBoardInfo())
        solution_file.write(board.printCarFuel() + "\n")

        solution_file.write("Runtime: " + str(ucs_solver.runtime) + " seconds\n")
        solution_file.write("Search Path Length: " + str(len(ucs_solver.searchPath)) + " states\n")
        solution_file.write("Solution Path Length: " + str(len(ucs_solver.solutionPath)) + " moves\n")
        solution_file.write("Solution Path: " + " No solutions! " + "\n\n")
        solution_file.close()
        sheet.cell(row=index*9 + 1 + 1, column=4).value = "N/A"
        sheet.cell(row=index*9 + 1 + 1, column=5).value = len(ucs_solver.searchPath)
        sheet.cell(row=index*9 + 1 + 1, column=6).value = ucs_solver.runtime

    search_file.write(ucs_solver.get_search_path())
    search_file.close()

    for h in range(4):
        sheet.cell(row=index*9 + 1 + 2 + h, column=1).value = index*9 + 1 + h + 1
        sheet.cell(row=index*9 + 1 + 2 + h, column=2).value = "A*"
        sheet.cell(row=index*9 + 1 + 2 + h, column=3).value = h+1
        solution_file = open(output_directory+"a-h" + str(h+1) + "-sol-" + str(index+1), 'w')
        search_file = open(output_directory + "a-h" + str(h+1) + "-search-" + str(index+1), 'w')

        astar_solver = AStar(board)
        astar_solver.search_solution(h + 1)

        if astar_solver.solutionFound:
            solution_file.write("--------------------------------------------------------------------------------\n\n")
            solution_file.write("Initial Configuration: \n\n")
            solution_file.write(board.returnBoardInfo())
            solution_file.write(board.printCarFuel() + "\n")

            solution_file.write("Runtime: " + str(astar_solver.runtime) + " seconds\n")
            solution_file.write("Search Path Length: " + str(len(astar_solver.searchPath)) + " states\n")
            solution_file.write("Solution Path Length: " + str(len(astar_solver.solutionPath) - 1) + " moves\n")
            solution_file.write("Solution Path: " + astar_solver.get_solution_moves() + "\n\n")
            solution_file.write(astar_solver.get_board_line() + "\n\n")

            solution_file.write("Final Configuration: \n\n")

            solution_file.write(astar_solver.solutionPath[0].board.returnBoardInfo())
            solution_file.close()
            sheet.cell(row=index*9 + 1 + 2 + h, column=4).value = len(astar_solver.solutionPath) - 1
            sheet.cell(row=index*9 + 1 + 2 + h, column=5).value = len(astar_solver.searchPath)
            sheet.cell(row=index*9 + 1 + 2 + h, column=6).value = astar_solver.runtime
        else:
            solution_file.write("--------------------------------------------------------------------------------\n\n")
            solution_file.write("Initial Configuration: \n\n")
            solution_file.write(board.returnBoardInfo())
            solution_file.write(board.printCarFuel() + "\n")

            solution_file.write("Runtime: " + str(astar_solver.runtime) + " seconds\n")
            solution_file.write("Search Path Length: " + str(len(astar_solver.searchPath)) + " states\n")
            solution_file.write("Solution Path Length: " + str(len(astar_solver.solutionPath)) + " moves\n")
            solution_file.write("Solution Path: " + " No solutions! " + "\n\n")
            solution_file.close()
            sheet.cell(row=index*9 + 1 + 2 + h, column=4).value = "N/A"
            sheet.cell(row=index*9+ 1 + 2 + h, column=5).value = len(astar_solver.searchPath)
            sheet.cell(row=index*9 + 1 + 2 + h, column=6).value = astar_solver.runtime

        search_file.write(astar_solver.get_search_path())
        search_file.close()

    for h in range(4):
        sheet.cell(row=index*9 + 1 + 2 + 4 + h, column=1).value = index*9 + 5 + h + 1
        sheet.cell(row=index*9 + 1 + 2 + 4 + h, column=2).value = "GBFS"
        sheet.cell(row=index*9 + 1 + 2 + 4 + h, column=3).value = h + 1

        solution_file = open(output_directory + "gbfs-h" + str(h + 1) + "-sol-" + str(index + 1), 'w')
        search_file = open(output_directory + "gbfs-h" + str(h + 1) + "-search-" + str(index + 1), 'w')

        gbfs_solver = GBFS(board)
        gbfs_solver.search_solution(h + 1)

        if gbfs_solver.solutionFound:
            solution_file.write("--------------------------------------------------------------------------------\n\n")
            solution_file.write("Initial Configuration: \n\n")
            solution_file.write(board.returnBoardInfo())
            solution_file.write(board.printCarFuel() + "\n")

            solution_file.write("Runtime: " + str(gbfs_solver.runtime) + " seconds\n")
            solution_file.write("Search Path Length: " + str(len(gbfs_solver.searchPath)) + " states\n")
            solution_file.write("Solution Path Length: " + str(len(gbfs_solver.solutionPath) - 1) + " moves\n")
            solution_file.write("Solution Path: " + gbfs_solver.get_solution_moves() + "\n\n")
            solution_file.write(gbfs_solver.get_board_line() + "\n\n")

            solution_file.write("Final Configuration: \n\n")

            solution_file.write(gbfs_solver.solutionPath[0].board.returnBoardInfo())
            solution_file.close()
            sheet.cell(row=index*9 + 1 + 2 + 4 + h, column=4).value = len(gbfs_solver.solutionPath) - 1
            sheet.cell(row=index*9 + 1 + 2 + 4 + h, column=5).value = len(gbfs_solver.searchPath)
            sheet.cell(row=index*9 + 1 + 2 + 4 + h, column=6).value = gbfs_solver.runtime
        else:
            solution_file.write("--------------------------------------------------------------------------------\n\n")
            solution_file.write("Initial Configuration: \n\n")
            solution_file.write(board.returnBoardInfo())
            solution_file.write(board.printCarFuel() + "\n")

            solution_file.write("Runtime: " + str(gbfs_solver.runtime) + " seconds\n")
            solution_file.write("Search Path Length: " + str(len(gbfs_solver.searchPath)) + " states\n")
            solution_file.write("Solution Path Length: " + str(len(gbfs_solver.solutionPath)) + " moves\n")
            solution_file.write("Solution Path: " + " No solutions! " + "\n\n")
            solution_file.close()
            sheet.cell(row=index*9 + 1 + 2 + 4 + h, column=4).value = "N/A"
            sheet.cell(row=index*9 + 1 + 2 + 4 + h, column=5).value = len(gbfs_solver.searchPath)
            sheet.cell(row=index*9 + 1 + 2 + 4 + h, column=6).value = gbfs_solver.runtime

        search_file.write(gbfs_solver.get_search_path())
        search_file.close()




print("Total Program run time: ", time.time() - start_time)
workbook.save("50_puzzles_result.xlsx")

